from __future__ import annotations

from datetime import date
from hashlib import sha256
from io import BytesIO
import sqlite3

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.worksheet import Worksheet

DISPLAY_COLUMNS = ["日付", "曜日", "時間", "診療科名", "変更前医師", "変更内容", "備考"]
INTERNAL_COLUMNS = ["登録種別", "レコードID", "差分キー"]
EXCEL_COLUMNS = DISPLAY_COLUMNS
DIFF_NEW = "新規"
DIFF_UPDATED = "更新"
DIFF_UNCHANGED = "既出"
HIGHLIGHT_FILL = PatternFill(fill_type="solid", fgColor="FF00FFFF")
HEADER_FILL = PatternFill(fill_type="solid", fgColor="D9EAF7")
FONT_SIZE = 24
ROW_HEIGHT_PIXELS = 55
ROW_HEIGHT_POINTS = ROW_HEIGHT_PIXELS * 0.75
COLUMN_WIDTHS = {
    "A": 28.33,
    "B": 13.33,
    "C": 19.00,
    "D": 41.67,
    "E": 39.67,
    "F": 103.67,
    "G": 109.33,
}
THIN_BORDER = Border(
    left=Side(style="thin", color="808080"),
    right=Side(style="thin", color="808080"),
    top=Side(style="thin", color="808080"),
    bottom=Side(style="thin", color="808080"),
)


def normalize_cell(value: object) -> str:
    if value is None or pd.isna(value):
        return ""
    return str(value)


def build_row_hash(row: pd.Series) -> str:
    payload = "\x1f".join(normalize_cell(row.get(column)) for column in DISPLAY_COLUMNS)
    return sha256(payload.encode("utf-8")).hexdigest()


def ensure_report2_output_tables(conn: sqlite3.Connection) -> None:
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS T_Report2OutputHistory (
            OutputHistoryID INTEGER PRIMARY KEY,
            OutputMode TEXT NOT NULL,
            OutputStatus TEXT NOT NULL DEFAULT 'active',
            StartDate DATE NOT NULL,
            OutputBy TEXT,
            OutputDate DATE,
            FileName TEXT,
            RecordCount INTEGER DEFAULT 0,
            CreatedAt DATETIME DEFAULT (datetime('now', '+9 hours')),
            CancelledAt DATETIME,
            CancelledBy TEXT,
            CancelReason TEXT
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS T_Report2OutputHistoryDetail (
            OutputDetailID INTEGER PRIMARY KEY,
            OutputHistoryID INTEGER NOT NULL,
            DiffKey TEXT NOT NULL,
            TargetType TEXT NOT NULL,
            TargetID INTEGER NOT NULL,
            RowHash TEXT NOT NULL,
            DiffStatus TEXT,
            ReportDate DATE,
            Weekday TEXT,
            TimeSlot TEXT,
            ClinicalDepartmentName TEXT,
            BeforeDoctorName TEXT,
            ChangeDetail TEXT,
            Reason TEXT,
            CreatedAt DATETIME DEFAULT (datetime('now', '+9 hours')),
            FOREIGN KEY (OutputHistoryID)
                REFERENCES T_Report2OutputHistory(OutputHistoryID)
        )
        """
    )


def load_latest_official_snapshot(conn: sqlite3.Connection) -> dict[str, str]:
    ensure_report2_output_tables(conn)
    row = conn.execute(
        """
        SELECT OutputHistoryID
        FROM T_Report2OutputHistory
        WHERE OutputMode = 'official'
          AND OutputStatus = 'active'
        ORDER BY CreatedAt DESC, OutputHistoryID DESC
        LIMIT 1
        """
    ).fetchone()
    if row is None:
        return {}

    detail_rows = conn.execute(
        """
        SELECT DiffKey, RowHash
        FROM T_Report2OutputHistoryDetail
        WHERE OutputHistoryID = ?
        """,
        (row[0],),
    ).fetchall()
    return {detail_row[0]: detail_row[1] for detail_row in detail_rows}


def add_diff_status(df: pd.DataFrame, previous_snapshot: dict[str, str]) -> pd.DataFrame:
    result = df.copy()
    if result.empty:
        result["差分区分"] = []
        result["RowHash"] = []
        return result

    result["RowHash"] = result.apply(build_row_hash, axis=1)

    def classify(row: pd.Series) -> str:
        previous_hash = previous_snapshot.get(normalize_cell(row["差分キー"]))
        if previous_hash is None:
            return DIFF_NEW
        if previous_hash != row["RowHash"]:
            return DIFF_UPDATED
        return DIFF_UNCHANGED

    result["差分区分"] = result.apply(classify, axis=1)
    return result


def cm_to_inches(cm: float) -> float:
    return cm / 2.54


def apply_print_settings(ws: Worksheet) -> None:
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:G{ws.max_row}"
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 55
    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_options.horizontalCentered = True
    ws.print_title_rows = "1:3"
    ws.oddFooter.right.text = "&D"
    ws.page_margins.left = cm_to_inches(0.3)
    ws.page_margins.right = cm_to_inches(0.3)
    ws.page_margins.top = cm_to_inches(1)
    ws.page_margins.bottom = cm_to_inches(1)
    ws.page_margins.header = cm_to_inches(1.9)
    ws.page_margins.footer = cm_to_inches(0.6)


def build_report2_excel(df_with_diff: pd.DataFrame, output_mode_label: str) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "帳票②予定変更一覧"

    ws["A1"] = "担当医師変更連絡表"
    ws["A1"].font = Font(size=FONT_SIZE, bold=False)
    ws.append([])
    ws.append(EXCEL_COLUMNS)
    for cell in ws[3]:
        cell.fill = HEADER_FILL
        cell.font = Font(size=FONT_SIZE, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER

    diagonal_border = Border(
        left=Side(style="thin", color="808080"),
        right=Side(style="thin", color="808080"),
        top=Side(style="thin", color="808080"),
        bottom=Side(style="thin", color="808080"),
        diagonal=Side(style="thin", color="808080"),
        diagonalUp=True,
    )

    change_detail_col = EXCEL_COLUMNS.index("変更内容") + 1
    for _, row in df_with_diff.iterrows():
        ws.append([normalize_cell(row.get(column)) for column in EXCEL_COLUMNS])
        excel_row = ws.max_row
        is_changed = row["差分区分"] in {DIFF_NEW, DIFF_UPDATED}
        change_text = normalize_cell(row.get("変更内容"))
        reason_text = normalize_cell(row.get("備考"))
        is_rest_day = change_text == "休診"
        # 「通常通り」は現在の正しい状態を示すため、備考に「取消」があっても
        # 帳票上で無効行のように見える斜線は付けない。
        is_cancelled = "取消" in reason_text and "通常通り" not in change_text

        for cell in ws[excel_row]:
            cell.border = diagonal_border if is_cancelled else THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.font = Font(size=FONT_SIZE, bold=is_changed)
            if is_changed:
                cell.fill = HIGHLIGHT_FILL
        if is_rest_day and not is_changed:
            ws.cell(row=excel_row, column=change_detail_col).fill = HIGHLIGHT_FILL

    for column_letter, width in COLUMN_WIDTHS.items():
        ws.column_dimensions[column_letter].width = width

    for row_number in range(1, ws.max_row + 1):
        ws.row_dimensions[row_number].height = ROW_HEIGHT_POINTS

    apply_print_settings(ws)

    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def save_official_output_history(
    conn: sqlite3.Connection,
    df_with_diff: pd.DataFrame,
    *,
    start_date: date,
    output_by: str,
    output_date: date,
    file_name: str,
) -> int:
    ensure_report2_output_tables(conn)
    cursor = conn.execute(
        """
        INSERT INTO T_Report2OutputHistory (
            OutputMode, OutputStatus, StartDate, OutputBy, OutputDate, FileName, RecordCount
        )
        VALUES ('official', 'active', ?, ?, ?, ?, ?)
        """,
        (str(start_date), output_by or None, str(output_date), file_name, int(len(df_with_diff))),
    )
    output_history_id = int(cursor.lastrowid)
    rows = []
    for _, row in df_with_diff.iterrows():
        rows.append(
            (
                output_history_id,
                normalize_cell(row["差分キー"]),
                normalize_cell(row["登録種別"]),
                int(row["レコードID"]),
                normalize_cell(row["RowHash"]),
                normalize_cell(row["差分区分"]),
                normalize_cell(row["日付"]),
                normalize_cell(row["曜日"]),
                normalize_cell(row["時間"]),
                normalize_cell(row["診療科名"]),
                normalize_cell(row["変更前医師"]),
                normalize_cell(row["変更内容"]),
                normalize_cell(row["備考"]),
            )
        )
    conn.executemany(
        """
        INSERT INTO T_Report2OutputHistoryDetail (
            OutputHistoryID, DiffKey, TargetType, TargetID, RowHash, DiffStatus,
            ReportDate, Weekday, TimeSlot, ClinicalDepartmentName, BeforeDoctorName,
            ChangeDetail, Reason
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        rows,
    )
    conn.commit()
    return output_history_id
