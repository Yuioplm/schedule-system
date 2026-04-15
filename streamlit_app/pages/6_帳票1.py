import streamlit as st
import pandas as pd
from io import BytesIO
import re

from openpyxl import load_workbook
from openpyxl.utils.cell import coordinate_to_tuple

from scripts.settings import get_conn

from streamlit_app.sql_loader import load_sql


def apply_placeholder_mappings(worksheet, df: pd.DataFrame, year: int, month: int) -> None:
    def normalize_text(value) -> str:
        text = "" if value is None else str(value)
        return text.replace("\u3000", " ").strip().lower()

    def filter_by_value(source_df: pd.DataFrame, col_name: str, target: str) -> pd.DataFrame:
        normalized_target = normalize_text(target)
        if not normalized_target or normalized_target == "*":
            return source_df

        normalized_series = source_df[col_name].map(normalize_text)
        return source_df[normalized_series == normalized_target]

    merged_anchor_map = {}
    for merged_range in worksheet.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged_range.bounds
        anchor = (min_row, min_col)
        for row_num in range(min_row, max_row + 1):
            for col_num in range(min_col, max_col + 1):
                merged_anchor_map[(row_num, col_num)] = anchor

    def write_value(cell_ref: str, value) -> None:
        row_num, col_num = coordinate_to_tuple(cell_ref)
        anchor_row, anchor_col = merged_anchor_map.get((row_num, col_num), (row_num, col_num))
        worksheet.cell(row=anchor_row, column=anchor_col, value=value)

    # 使用可能な書式:
    # {{年}} / {{月}} / {{固定:文字列}}
    # {{曜日|診療科|時間|診察室}} 例: {{月|内科|午前|101}}
    # ※ 診療科・診察室はワイルドカード（*）可
    # 旧形式: {{列名#行番号}} も後方互換で許可
    token_pattern = re.compile(r"^\s*\{\{\s*(.+?)\s*\}\}\s*$")
    for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
        for cell in row:
            if not isinstance(cell.value, str):
                continue
            matched = token_pattern.match(cell.value)
            if not matched:
                continue

            token = matched.group(1).strip()
            replacement = ""
            if token == "年":
                replacement = year
            elif token == "月":
                replacement = month
            elif token.startswith("固定:"):
                replacement = token.split(":", 1)[1]
            elif "|" in token or "｜" in token:
                normalized_token = token.replace("｜", "|")
                parts = [part.strip() for part in normalized_token.split("|")]
                if len(parts) >= 3:
                    weekday, dept, time_slot = parts[:3]
                    room = parts[3] if len(parts) >= 4 else ""

                    if weekday in df.columns:
                        candidate_df = df.copy()
                        if "診療科" in candidate_df.columns:
                            candidate_df = filter_by_value(candidate_df, "診療科", dept)
                        if "時間" in candidate_df.columns:
                            candidate_df = filter_by_value(candidate_df, "時間", time_slot)

                        room_col = None
                        for col_name in ["診察室", "部屋番号", "部屋"]:
                            if col_name in candidate_df.columns:
                                room_col = col_name
                                break
                        if room_col:
                            candidate_df = filter_by_value(candidate_df, room_col, room)

                        if len(candidate_df) > 0:
                            src_value = candidate_df.iloc[0][weekday]
                            replacement = "" if pd.isna(src_value) else str(src_value)
            elif "#" in token:
                col_name, row_str = token.split("#", 1)
                col_name = col_name.strip()
                if col_name in df.columns and row_str.isdigit():
                    row_idx = int(row_str) - 1
                    if 0 <= row_idx < len(df):
                        src_value = df.iloc[row_idx][col_name]
                        replacement = "" if pd.isna(src_value) else str(src_value)
            elif token in df.columns and len(df) > 0:
                src_value = df.iloc[0][token]
                replacement = "" if pd.isna(src_value) else str(src_value)

            write_value(cell.coordinate, replacement)


st.set_page_config(layout="wide")
st.title("帳票① 外来担当医表")

conn = get_conn()

col1, col2 = st.columns(2)
with col1:
    year = st.number_input("年", min_value=2020, max_value=2100, value=2026, step=1)
with col2:
    month = st.number_input("月", min_value=1, max_value=12, value=4, step=1)

subcategory = st.radio(
    "小カテゴリ",
    ["院内用", "外部用"],
    horizontal=True,
    help="院内用は既存仕様、外部用は医師名をフルネーム・週表記を「第1・3...」で出力します。",
)

sql_name = "Report1_pivot.sql" if subcategory == "院内用" else "Report1_pivot_external.sql"

target_month = f"{int(year)}-{int(month):02d}"
query = load_sql(sql_name)
df = pd.read_sql(query, conn, params={"target_month": target_month})

if df.empty:
    st.warning("対象月のデータがありません")
else:
    display_mode = st.radio(
        "画面表示形式",
        ["改行表示", "区切り表示（ / ）"],
        horizontal=True,
    )

    display_df = df.copy()

    day_cols = ["月", "火", "水", "木", "金", "土"]
    for col in day_cols:
        if col in display_df.columns:
            display_df[col] = display_df[col].fillna("")

    if display_mode == "区切り表示（ / ）":
        for col in day_cols:
            if col in display_df.columns:
                display_df[col] = display_df[col].str.replace("\n", " / ")
        st.dataframe(display_df, use_container_width=True)
    else:
        styled_df = display_df.style.set_properties(
            subset=day_cols,
            **{"white-space": "pre-wrap"}
        )
        st.dataframe(styled_df, use_container_width=True)

    csv = df.to_csv(index=False).encode("utf-8-sig")
    st.download_button(
        label="Excelダウンロード",
        data=csv,
        file_name=f"帳票①_外来担当医表_{subcategory}_{target_month}.csv",
        mime="text/csv",
    )

    st.markdown("#### テンプレート反映設定")
    st.info(
        "Excelテンプレート（.xlsx）にプレースホルダ（例: {{月|内科|午前|10}}）を入力しておくと、"
        "対応するデータが反映されたExcelをダウンロードできます。"
    )
    st.caption(
        "プレースホルダ書式: {{年}} / {{月}} / {{固定:文字列}} / {{曜日|診療科|時間|診察室}} "
        "（例: {{月|内科|午前|10}}, {{火|*|午後|*}}, {{固定:休診}}）※ `|` と `｜` の両方可"
    )
    with st.expander("操作ガイド", expanded=True):
        st.markdown(
            """
            1. Excelテンプレートの値を入れたいセルに、プレースホルダを直接入力します。  
            2. 例: `{{年}}`, `{{月}}`, `{{月|内科|午前|101}}`, `{{火|*|午後|*}}`, `{{固定:休診}}`。  
            3. 画面でテンプレートをアップロードし、**テンプレート反映版Excelダウンロード** を押します。  
            4. 出力結果を確認し、必要ならテンプレート側のプレースホルダを修正します。  
            5. `*` はワイルドカードです（例: 診療科を問わない `{{月|*|午前|101}}`、部屋番号を問わない `{{月|内科|午前|*}}`）。  
            """
        )

    template_file = st.file_uploader(
        "帳票①Excelテンプレート（.xlsx）",
        type=["xlsx"],
        help=(
            "テンプレートをアップロードすると、見た目（罫線・結合セル）を維持したまま、"
            "担当医表の値を該当セルへ流し込んだExcelを出力します。"
        ),
    )

    if template_file is not None:
        template_bytes = template_file.getvalue()
        try:
            workbook = load_workbook(BytesIO(template_bytes))
            worksheet = workbook.active
            apply_placeholder_mappings(
                worksheet=worksheet,
                df=df,
                year=int(year),
                month=int(month),
            )
            output = BytesIO()
            workbook.save(output)
            filled_excel = output.getvalue()
            st.download_button(
                label="テンプレート反映版Excelダウンロード",
                data=filled_excel,
                file_name=f"帳票①_外来担当医表_{subcategory}_{target_month}_template.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        except Exception as exc:
            st.error(f"テンプレートExcelへの反映に失敗しました: {exc}")
