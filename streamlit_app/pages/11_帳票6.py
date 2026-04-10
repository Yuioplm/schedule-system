import calendar
from datetime import datetime
from io import BytesIO
from pathlib import Path
import re
import sys

import pandas as pd
import streamlit as st
from openpyxl.cell.cell import MergedCell
from openpyxl import load_workbook

from scripts.settings import get_conn

sys.path.append(str(Path(__file__).resolve().parents[2]))

from streamlit_app.sql_loader import load_sql


def classify_period(timeslot_name: str) -> str | None:
    if timeslot_name is None:
        return None

    name = str(timeslot_name).strip().lower()
    if not name:
        return None

    if "午前" in name or "am" in name or "morning" in name or "前" in name:
        return "AM"
    if "午後" in name or "pm" in name or "afternoon" in name or "後" in name:
        return "PM"
    return None


def fetch_eligible_doctors(conn, start_date: str, end_date: str) -> pd.DataFrame:
    query = load_sql("Report6_eligible_doctors.sql")
    params = {"start_date": start_date, "end_date": end_date}
    try:
        return pd.read_sql(query, conn, params=params)
    except TypeError:
        # 環境差異により cursor.description が None となる場合のフォールバック
        cursor = conn.execute(query, params)
        rows = cursor.fetchall()
        if cursor.description is None:
            return pd.DataFrame(columns=["DoctorID", "DoctorName", "所属"])
        columns = [col_desc[0] for col_desc in cursor.description]
        return pd.DataFrame(rows, columns=columns)


def fetch_report6_rows(conn, start_date: str, end_date: str) -> pd.DataFrame:
    query = load_sql("Report6_daily_status.sql")
    return pd.read_sql(
        query,
        conn,
        params={"start_date": start_date, "end_date": end_date},
    )


def build_report_dataframe(
    doctor_id: int,
    year: int,
    month: int,
    source_df: pd.DataFrame,
) -> pd.DataFrame:
    days_in_month = calendar.monthrange(year, month)[1]
    date_index = pd.date_range(f"{year}-{month:02d}-01", periods=days_in_month, freq="D")
    weekday_map = ["月", "火", "水", "木", "金", "土", "日"]

    day_df = pd.DataFrame({
        "CalendarDate": date_index.strftime("%Y-%m-%d"),
        "日": list(range(1, days_in_month + 1)),
        "曜日": [weekday_map[d.weekday()] for d in date_index],
        "AM勤務": [None] * days_in_month,
        "PM勤務": [None] * days_in_month,
        "備考": [None] * days_in_month,
    })

    doctor_rows = source_df[source_df["DoctorID"] == doctor_id].copy()
    doc_actual = doctor_rows[doctor_rows["RowType"] == "ACTUAL"].copy()
    doc_rest = doctor_rows[doctor_rows["RowType"] == "REST"].copy()
    doc_temp = doctor_rows[doctor_rows["RowType"] == "TEMP"].copy()

    for source_df in [doc_actual, doc_rest, doc_temp]:
        source_df["Period"] = source_df["TimeSlotName"].map(classify_period)

    worked = (
        doc_actual.dropna(subset=["Period"])
        .groupby(["CalendarDate", "Period"], as_index=False)
        .size()
    )
    rest = (
        doc_rest.dropna(subset=["Period"])
        .groupby(["CalendarDate", "Period"], as_index=False)
        .size()
    )
    temporary = (
        doc_temp.dropna(subset=["Period"])
        .groupby(["CalendarDate", "Period"], as_index=False)
        .size()
    )

    status_map: dict[tuple[str, str], str] = {}
    for _, row in rest.iterrows():
        status_map[(row["CalendarDate"], row["Period"])] = "休み"
    for _, row in worked.iterrows():
        status_map[(row["CalendarDate"], row["Period"])] = "〇"

    temp_map: dict[str, list[str]] = {}
    for _, row in temporary.iterrows():
        date_key = row["CalendarDate"]
        period = row["Period"]
        if period == "AM":
            label = "AM臨時出勤"
        elif period == "PM":
            label = "PM臨時出勤"
        else:
            continue
        temp_map.setdefault(date_key, [])
        if label not in temp_map[date_key]:
            temp_map[date_key].append(label)

    for idx, row in day_df.iterrows():
        cal_date = row["CalendarDate"]
        day_df.at[idx, "AM勤務"] = status_map.get((cal_date, "AM"))
        day_df.at[idx, "PM勤務"] = status_map.get((cal_date, "PM"))
        remarks = temp_map.get(cal_date, [])
        day_df.at[idx, "備考"] = " / ".join(remarks) if remarks else None

    return day_df[["日", "曜日", "AM勤務", "PM勤務", "備考"]]


def write_doctor_sheet(
    worksheet,
    report_df: pd.DataFrame,
    doctor_name: str,
    doctor_id: int,
    department: str,
    year: int,
    month: int,
) -> None:
    def set_cell_value_safe(row_no: int, col_no: int, value) -> None:
        cell = worksheet.cell(row=row_no, column=col_no)
        if isinstance(cell, MergedCell):
            # 結合セルの先頭セル以外は書き込み不可のためスキップ
            return
        cell.value = value

    day_column_names = {"曜日", "AM勤務", "PM勤務", "備考"}
    day_value_map = {
        (int(row["日"]), col_name): row[col_name]
        for _, row in report_df.iterrows()
        for col_name in day_column_names
    }

    day_pattern = re.compile(r"\{\{(\d{1,2})\|([^}]+)\}\}")

    def replace_day_placeholder(text: str) -> str:
        def _repl(match: re.Match[str]) -> str:
            day = int(match.group(1))
            col_name = match.group(2).strip()
            if not (1 <= day <= calendar.monthrange(year, month)[1]):
                return ""
            if col_name not in day_column_names:
                return ""
            value = day_value_map.get((day, col_name))
            return "" if pd.isna(value) else str(value)

        return day_pattern.sub(_repl, text)

    marker_row = None
    marker_col = None
    for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
        for cell in row:
            if isinstance(cell, MergedCell):
                continue
            if isinstance(cell.value, str):
                raw = cell.value.strip()
                if raw == "{{明細開始}}":
                    marker_row, marker_col = cell.row, cell.column
                    cell.value = None
                else:
                    replaced = (
                        raw.replace("{{医師名}}", doctor_name)
                        .replace("{{ID}}", str(doctor_id))
                        .replace("{{所属}}", department)
                        .replace("{{年}}", str(year))
                        .replace("{{月}}", str(month))
                    )
                    cell.value = replace_day_placeholder(replaced)

    if marker_row is None or marker_col is None:
        # {{明細開始}} が無い場合は、プレースホルダ置換のみ行い
        # 明細表の自動書き込みはしない（テンプレート指定位置を優先）
        return

    for i, data_row in enumerate(report_df.itertuples(index=False), start=0):
        row_no = marker_row + i
        set_cell_value_safe(row_no, marker_col, data_row[0])
        set_cell_value_safe(row_no, marker_col + 1, data_row[1])
        set_cell_value_safe(row_no, marker_col + 2, data_row[2])
        set_cell_value_safe(row_no, marker_col + 3, data_row[3])
        set_cell_value_safe(row_no, marker_col + 4, data_row[4])


st.set_page_config(layout="wide")
st.title("帳票⑥ 非常勤医師勤務報告書")

conn = get_conn()

this_year = datetime.now().year
this_month = datetime.now().month

col1, col2 = st.columns(2)
with col1:
    year = st.number_input("年", min_value=2020, max_value=2100, value=this_year, step=1)
with col2:
    month = st.number_input("月", min_value=1, max_value=12, value=this_month, step=1)

start_date = f"{int(year)}-{int(month):02d}-01"
end_date = f"{int(year)}-{int(month):02d}-{calendar.monthrange(int(year), int(month))[1]:02d}"

doctor_df = fetch_eligible_doctors(conn, start_date, end_date)
if doctor_df.empty:
    st.warning("選択月に外来予定または実績がある非常勤医師がいません")
    st.stop()

report6_source_df = fetch_report6_rows(conn, start_date, end_date)

doctor_df["所属"] = doctor_df["所属"].fillna("")
department_options = ["(全て)"] + sorted([x for x in doctor_df["所属"].unique().tolist() if x])
selected_department = st.selectbox("所属", department_options)

filtered_doctor_df = doctor_df.copy()
if selected_department != "(全て)":
    filtered_doctor_df = filtered_doctor_df[filtered_doctor_df["所属"] == selected_department]

if filtered_doctor_df.empty:
    st.warning("選択した所属に該当する医師がいません")
    st.stop()

doctor_options = filtered_doctor_df["DoctorID"].astype(int).tolist()
selected_doctor_id = st.selectbox(
    "医師名",
    doctor_options,
    format_func=lambda doc_id: (
        f"{filtered_doctor_df.loc[filtered_doctor_df['DoctorID'] == doc_id, 'DoctorName'].iloc[0]} "
        f"(ID:{doc_id}, 所属:{filtered_doctor_df.loc[filtered_doctor_df['DoctorID'] == doc_id, '所属'].iloc[0] or '-'})"
    ),
)
selected_doctor_row = filtered_doctor_df.loc[filtered_doctor_df["DoctorID"] == selected_doctor_id].iloc[0]
selected_doctor_name = str(selected_doctor_row["DoctorName"])

preview_df = build_report_dataframe(
    doctor_id=selected_doctor_id,
    year=int(year),
    month=int(month),
    source_df=report6_source_df,
)

st.markdown("#### 個人別プレビュー")
st.dataframe(preview_df, use_container_width=True)

csv = preview_df.to_csv(index=False).encode("utf-8-sig")
st.download_button(
    label="プレビューCSVダウンロード",
    data=csv,
    file_name=f"帳票⑥_{year}_{int(month):02d}_{selected_doctor_name}.csv",
    mime="text/csv",
)

st.markdown("#### Excelテンプレート（個人別シート出力）")
st.caption("テンプレートをアップロードすると、選択月に外来予定または実績がある非常勤医師ごとのシート（シート名：医師名）を作成してExcelを出力します。")
st.caption("テンプレート内プレースホルダ: {{医師名}}, {{ID}}, {{所属}}, {{年}}, {{月}}, {{明細開始}}（明細表を出力する場合に必須）")
st.caption("日別カラム指定プレースホルダ: {{1|曜日}}, {{5|AM勤務}}, {{7|PM勤務}}, {{12|備考}}")

template_file = st.file_uploader("帳票⑥Excelテンプレート（.xlsx）", type=["xlsx"])

if template_file is not None:
    try:
        wb = load_workbook(BytesIO(template_file.getvalue()))
        base_ws = wb.active

        doctor_rows = doctor_df.to_dict("records")
        if not doctor_rows:
            st.warning("対象の非常勤医師がいないため、Excelを作成できません")
        else:
            existing_names = set()
            sheet_plan: list[tuple[dict, object]] = []

            # 先に全シートを「未加工テンプレート」から複製しておく
            # （1件目の書き込み後に複製すると、以降のシートが1件目内容を引き継いでしまうため）
            first = True
            for doctor in doctor_rows:
                doctor_id_raw = int(doctor["DoctorID"])
                doc_name = str(doctor["DoctorName"]) if doctor["DoctorName"] else f"Doctor_{doctor_id_raw}"

                if first:
                    ws = base_ws
                    first = False
                else:
                    ws = wb.copy_worksheet(base_ws)

                sheet_name = doc_name[:31]
                suffix = 1
                while sheet_name in existing_names:
                    tail = f"_{suffix}"
                    sheet_name = f"{doc_name[:31-len(tail)]}{tail}"
                    suffix += 1
                ws.title = sheet_name
                existing_names.add(sheet_name)
                sheet_plan.append((doctor, ws))

            for doctor, ws in sheet_plan:
                doc_id = int(doctor["DoctorID"])
                doc_name = str(doctor["DoctorName"]) if doctor["DoctorName"] else f"Doctor_{doc_id}"
                doc_department = str(doctor["所属"]) if doctor["所属"] else ""
                doc_report_df = build_report_dataframe(
                    doctor_id=doc_id,
                    year=int(year),
                    month=int(month),
                    source_df=report6_source_df,
                )

                write_doctor_sheet(
                    ws,
                    doc_report_df,
                    doc_name,
                    doc_id,
                    doc_department,
                    int(year),
                    int(month),
                )

            output = BytesIO()
            wb.save(output)
            st.download_button(
                label="個人別Excelダウンロード",
                data=output.getvalue(),
                file_name=f"帳票⑥_非常勤医師勤務報告書_{year}_{int(month):02d}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
    except Exception as exc:
        st.error(f"Excel生成に失敗しました: {exc}")
