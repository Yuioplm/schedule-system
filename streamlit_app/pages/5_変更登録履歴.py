import streamlit as st
import pandas as pd
from io import BytesIO
import re
import zipfile
import xml.etree.ElementTree as ET
from xml.sax.saxutils import escape
from time import perf_counter

from openpyxl import load_workbook
from openpyxl.utils.cell import coordinate_to_tuple

from scripts.settings import get_conn
from streamlit_app.log_events import log_event, log_page_open
from streamlit_app.sql_loader import load_sql


def build_history_placeholder_replacements(template_bytes: bytes, df: pd.DataFrame) -> tuple[str, dict[str, str]]:
    workbook = load_workbook(BytesIO(template_bytes))
    worksheet = workbook.active

    def normalize_text(value) -> str:
        return "" if value is None else str(value).replace("\u3000", " ").strip()

    merged_anchor_map = {}
    for merged_range in worksheet.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged_range.bounds
        anchor = (min_row, min_col)
        for row_num in range(min_row, max_row + 1):
            for col_num in range(min_col, max_col + 1):
                merged_anchor_map[(row_num, col_num)] = anchor

    replacements: dict[str, str] = {}

    def write_value(cell_ref: str, value) -> None:
        row_num, col_num = coordinate_to_tuple(cell_ref)
        anchor_row, anchor_col = merged_anchor_map.get((row_num, col_num), (row_num, col_num))
        anchor_ref = worksheet.cell(row=anchor_row, column=anchor_col).coordinate
        replacements[anchor_ref] = value

    token_pattern = re.compile(r"^\s*\{\{\s*(.+?)\s*\}\}\s*$")
    for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
        for cell in row:
            if not isinstance(cell.value, str):
                continue
            matched = token_pattern.match(cell.value)
            if not matched:
                continue

            token = matched.group(1).strip()
            replacement = ""
            if token.startswith("固定:"):
                replacement = token.split(":", 1)[1]
            elif "#" in token:
                col_name, row_str = token.split("#", 1)
                col_name = normalize_text(col_name)
                if col_name in df.columns and row_str.isdigit():
                    row_idx = int(row_str) - 1
                    if 0 <= row_idx < len(df):
                        src_value = df.iloc[row_idx][col_name]
                        replacement = "" if pd.isna(src_value) else str(src_value)
            elif token in df.columns and len(df) > 0:
                src_value = df.iloc[0][token]
                replacement = "" if pd.isna(src_value) else str(src_value)

            write_value(cell.coordinate, replacement)

    return worksheet.title, replacements


def resolve_sheet_path_by_title(template_bytes: bytes, sheet_title: str) -> str:
    ns_main = {"main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    ns_rel = {"rel": "http://schemas.openxmlformats.org/officeDocument/2006/relationships"}
    ns_pkg = {"pkg": "http://schemas.openxmlformats.org/package/2006/relationships"}

    with zipfile.ZipFile(BytesIO(template_bytes), "r") as source_zip:
        workbook_root = ET.fromstring(source_zip.read("xl/workbook.xml"))
        relation_root = ET.fromstring(source_zip.read("xl/_rels/workbook.xml.rels"))

    sheet_rel_id = None
    for sheet in workbook_root.findall(".//main:sheets/main:sheet", ns_main):
        if sheet.attrib.get("name") == sheet_title:
            sheet_rel_id = sheet.attrib.get(f"{{{ns_rel['rel']}}}id")
            break

    if not sheet_rel_id:
        raise ValueError(f"対象シートが見つかりません: {sheet_title}")

    for rel in relation_root.findall(".//pkg:Relationship", ns_pkg):
        if rel.attrib.get("Id") != sheet_rel_id:
            continue
        target = rel.attrib.get("Target", "")
        target = target.lstrip("/")
        if target.startswith("xl/"):
            return target
        return f"xl/{target}"

    raise ValueError(f"対象シートの参照先が見つかりません: {sheet_title}")


def apply_replacements_preserve_media(template_bytes: bytes, sheet_title: str, replacements: dict[str, str]) -> bytes:
    if not replacements:
        return template_bytes

    sheet_path = resolve_sheet_path_by_title(template_bytes=template_bytes, sheet_title=sheet_title)

    source_buffer = BytesIO(template_bytes)
    output_buffer = BytesIO()

    with zipfile.ZipFile(source_buffer, "r") as source_zip, zipfile.ZipFile(output_buffer, "w") as output_zip:
        for item in source_zip.infolist():
            file_bytes = source_zip.read(item.filename)
            if item.filename == sheet_path:
                sheet_xml = file_bytes.decode("utf-8")
                for cell_ref, replacement in replacements.items():
                    sanitized_text = sanitize_xml_text(replacement)
                    sheet_xml, _ = replace_cell_value_with_inline_string(
                        sheet_xml=sheet_xml,
                        cell_ref=cell_ref,
                        text=sanitized_text,
                    )
                file_bytes = sheet_xml.encode("utf-8")

            output_zip.writestr(item, file_bytes)

    return output_buffer.getvalue()


def sanitize_xml_text(value: str) -> str:
    if not value:
        return ""

    # XML 1.0で許可されない文字を除去（タブ/改行/復帰は許可）
    value = re.sub(r"[\x00-\x08\x0B\x0C\x0E-\x1F]", "", value)
    value = re.sub(r"[\uD800-\uDFFF\uFFFE\uFFFF]", "", value)
    return value


def replace_cell_value_with_inline_string(sheet_xml: str, cell_ref: str, text: str) -> tuple[str, bool]:
    cell_pattern = re.compile(
        rf'(<c\b[^>]*\br="{re.escape(cell_ref)}"[^>]*)(?:\s*/>|>.*?</c>)',
        flags=re.DOTALL,
    )

    escaped_text = escape(text)
    text_attrs = ' xml:space="preserve"' if text[:1].isspace() or text[-1:].isspace() else ""

    def repl(match: re.Match) -> str:
        cell_start = re.sub(r'\s+t="[^"]*"', "", match.group(1))
        return f'{cell_start} t="inlineStr"><is><t{text_attrs}>{escaped_text}</t></is></c>'

    replaced_xml, count = cell_pattern.subn(repl, sheet_xml, count=1)
    return replaced_xml, count > 0


def save_output_history(conn, target_df: pd.DataFrame, output_by: str, output_date: str) -> None:
    if target_df.empty:
        return
    target_id_col = "レコードID" if "レコードID" in target_df.columns else "登録ID"
    if target_id_col not in target_df.columns:
        raise KeyError("レコードID / 登録ID の列が見つかりません。")
    rows = [
        (row["登録種別"], int(row[target_id_col]), output_by if output_by else None, output_date)
        for _, row in target_df.iterrows()
    ]
    conn.executemany(
        """
        INSERT INTO T_ChangeNoticeOutputHistory (
            TargetType, TargetID, OutputBy, OutputDate
        )
        VALUES (?, ?, ?, ?)
        """,
        rows,
    )
    conn.commit()


st.title("変更登録履歴検索")
log_page_open("変更登録履歴検索")
conn = get_conn()

st.caption("予定変更入力・臨時外来登録の入力内容を、非表示設定を含めて確認できます。")

COLUMN_LABEL_RENAMES = {
    "SlotID": "枠ID",
    "WeekPattern": "週番号",
    "診療科": "診療科名",
    "専門外来名": "帳票①用専門外来名",
    "帳票①表示名（任意）": "帳票①用医師表示名",
    "レコードID": "登録ID",
    "医師": "変更後医師名",
    "部屋": "変更後部屋",
    "帳票➁変更前": "変更前医師表示名",
    "変更種別": "変更種別名",
    "医師ID": "変更後医師ID",
    "時間帯ID": "臨時外来時間帯ID",
    "診療科ID": "臨時外来診療科ID",
    "EmploymentType": "勤務形態",
    "DoctorName": "医師名",
}


def _sync_end_date_from_start(start_key: str, end_key: str) -> None:
    st.session_state[end_key] = st.session_state[start_key]


if "change_history_date_from" not in st.session_state:
    st.session_state["change_history_date_from"] = pd.Timestamp.today().date()
if "change_history_date_to" not in st.session_state:
    st.session_state["change_history_date_to"] = st.session_state["change_history_date_from"]

col1, col2 = st.columns(2)
with col1:
    date_from = st.date_input(
        "開始日",
        key="change_history_date_from",
        on_change=_sync_end_date_from_start,
        args=("change_history_date_from", "change_history_date_to"),
    )
with col2:
    date_to = st.date_input("終了日", key="change_history_date_to")

show_inactive = st.checkbox("無効化済み(ActiveFlag=0)も表示", value=False)

query = load_sql("ChangeHistory_search.sql")

request_id = log_event(
    "search_execute",
    "変更登録履歴検索",
    date_from=str(date_from),
    date_to=str(date_to),
    show_inactive=show_inactive,
)
search_started_at = perf_counter()
try:
    result_df = pd.read_sql(
        query,
        conn,
        params=[str(date_from), str(date_to), str(date_from), str(date_to), 1 if show_inactive else 0],
    )
except Exception as exc:
    log_event(
        "search_failed",
        "変更登録履歴検索",
        request_id=request_id,
        error=type(exc).__name__,
    )
    raise

elapsed_ms = int((perf_counter() - search_started_at) * 1000)
log_event(
    "search_result",
    "変更登録履歴検索",
    request_id=request_id,
    result_count=len(result_df),
    elapsed_ms=elapsed_ms,
)
if not result_df.empty:
    result_df["日付"] = pd.to_datetime(result_df["日付"])
    result_df["曜日"] = result_df["日付"].dt.day_name().map(
        {
            "Monday": "月",
            "Tuesday": "火",
            "Wednesday": "水",
            "Thursday": "木",
            "Friday": "金",
            "Saturday": "土",
            "Sunday": "日",
        }
    )
    insert_at = result_df.columns.get_loc("日付") + 1
    result_df.insert(insert_at, "曜日", result_df.pop("曜日"))
    result_df["日付"] = result_df["日付"].dt.strftime("%Y-%m-%d")

st.subheader("検索結果")
if result_df.empty:
    st.info("該当データがありません")
else:
    display_columns = [col for col in result_df.columns if col != "編集日付"]
    display_result_df = result_df[display_columns].copy().rename(columns=COLUMN_LABEL_RENAMES)

    preferred_order = [
        "登録種別",
        "登録ID",
        "日付",
        "曜日",
        "枠ID",
        "時間帯",
        "診療科名",
        "変更前医師表示名",
        "変更種別ID",
        "変更種別名",
        "変更後部屋",
        "変更後医師ID",
        "変更後医師名",
        "変更内容",
        "備考",
        "臨時外来時間帯ID",
        "臨時外来診療科ID",
        "帳票②表示",
        "ActiveFlag",
        "登録者",
        "登録日時",
        "変更届出力者",
        "変更届出力日",
    ]
    remaining_cols = [col for col in display_result_df.columns if col not in preferred_order]
    display_result_df = display_result_df[[col for col in preferred_order if col in display_result_df.columns] + remaining_cols]
    st.dataframe(display_result_df, use_container_width=True)
    csv = display_result_df.to_csv(index=False).encode("utf-8-sig")
    st.download_button(
        label="CSVダウンロード",
        data=csv,
        file_name="変更登録履歴.csv",
        mime="text/csv",
    )

    st.markdown("---")
    st.subheader("登録内容の編集")

    edit_df = result_df.copy()
    edit_df["選択表示"] = edit_df.apply(
        lambda r: f"{r['登録種別']} / {r['日付']} / ID:{int(r['レコードID'])} / {r['変更種別'] or '-'}",
        axis=1,
    )

    selected_label = st.selectbox("編集対象", edit_df["選択表示"].tolist())
    selected_row = edit_df.loc[edit_df["選択表示"] == selected_label].iloc[0]

    is_visible_report2 = selected_row["帳票②表示"] == "表示"
    is_normal_change = selected_row["登録種別"] == "通常枠変更"

    change_type_df = pd.read_sql(
        """
        SELECT ChangeTypeID, ChangeTypeName
        FROM M_ScheduleChangeType
        WHERE ActiveFlag = 1
        ORDER BY ChangeTypeID
        """,
        conn,
    )
    doctor_df = pd.read_sql(
        """
        SELECT DoctorID, DoctorName
        FROM M_Doctor
        WHERE ActiveFlag = 1
        ORDER BY DoctorID
        """,
        conn,
    )
    timeslot_df = pd.read_sql(
        """
        SELECT TimeSlotID, TimeSlotName
        FROM M_TimeSlot
        ORDER BY TimeSlotID
        """,
        conn,
    )
    clin_dept_df = pd.read_sql(
        """
        SELECT ClinDeptID, ClinDeptName
        FROM M_ClinicalDepartment
        WHERE ActiveFlag = 1
        ORDER BY ClinDeptID
        """,
        conn,
    )

    with st.form("history_edit_form"):
        if is_normal_change:
            current_change_type = selected_row["変更種別ID"]
            change_type_options = change_type_df["ChangeTypeID"].astype(int).tolist()
            change_type_index = 0
            if pd.notna(current_change_type) and int(current_change_type) in change_type_options:
                change_type_index = change_type_options.index(int(current_change_type))
            edit_change_type_id = st.selectbox(
                "変更種別名",
                change_type_options,
                index=change_type_index,
                format_func=lambda x: f"{x}: {change_type_df.loc[change_type_df['ChangeTypeID'] == x, 'ChangeTypeName'].iloc[0]}",
            )

            current_doctor = selected_row["医師ID"]
            doctor_options = [None] + doctor_df["DoctorID"].astype(int).tolist()
            doctor_index = 0
            if pd.notna(current_doctor) and int(current_doctor) in doctor_options:
                doctor_index = doctor_options.index(int(current_doctor))
            edit_doctor_id = st.selectbox(
                "代診医（任意）",
                doctor_options,
                index=doctor_index,
                format_func=lambda x: "未設定" if x is None else f"{x}: {doctor_df.loc[doctor_df['DoctorID'] == x, 'DoctorName'].iloc[0]}",
            )
            edit_changed_by = st.text_input("ChangedBy", value=selected_row["登録者"] or "")
        else:
            edit_date = st.date_input("日付", value=pd.to_datetime(selected_row["編集日付"]).date())

            current_timeslot = int(selected_row["時間帯ID"]) if pd.notna(selected_row["時間帯ID"]) else None
            timeslot_options = timeslot_df["TimeSlotID"].astype(int).tolist()
            timeslot_index = 0
            if current_timeslot in timeslot_options:
                timeslot_index = timeslot_options.index(current_timeslot)
            edit_timeslot_id = st.selectbox(
                "臨時外来時間帯ID",
                timeslot_options,
                index=timeslot_index,
                format_func=lambda x: f"{x}: {timeslot_df.loc[timeslot_df['TimeSlotID'] == x, 'TimeSlotName'].iloc[0]}",
            )

            current_dept = int(selected_row["診療科ID"]) if pd.notna(selected_row["診療科ID"]) else None
            dept_options = clin_dept_df["ClinDeptID"].astype(int).tolist()
            dept_index = 0
            if current_dept in dept_options:
                dept_index = dept_options.index(current_dept)
            edit_dept_id = st.selectbox(
                "臨時外来診療科ID",
                dept_options,
                index=dept_index,
                format_func=lambda x: f"{x}: {clin_dept_df.loc[clin_dept_df['ClinDeptID'] == x, 'ClinDeptName'].iloc[0]}",
            )

            current_doctor = selected_row["医師ID"]
            doctor_options = [None] + doctor_df["DoctorID"].astype(int).tolist()
            doctor_index = 0
            if pd.notna(current_doctor) and int(current_doctor) in doctor_options:
                doctor_index = doctor_options.index(int(current_doctor))
            edit_doctor_id = st.selectbox(
                "変更後医師名（任意）",
                doctor_options,
                index=doctor_index,
                format_func=lambda x: "未設定" if x is None else f"{x}: {doctor_df.loc[doctor_df['DoctorID'] == x, 'DoctorName'].iloc[0]}",
            )

            edit_room = st.text_input("変更後部屋", value=selected_row["部屋"] or "")
            edit_rpt2_before_doctor = st.text_input("変更前医師表示名（任意）", value=selected_row["帳票➁変更前"] or "")

        edit_detail = st.text_area("変更内容", value=selected_row["変更内容"] or "")
        edit_reason = st.text_area("備考", value=selected_row["備考"] or "")
        edit_visible_report2 = st.checkbox("予定変更一覧に表示", value=is_visible_report2)
        edit_active = st.checkbox("有効", value=bool(selected_row["ActiveFlag"]))

        submitted_edit = st.form_submit_button("更新")

        if submitted_edit:
            update_request_id = log_event(
                "update_start",
                "変更登録履歴検索",
                operation="edit_history",
                target_type=selected_row["登録種別"],
                target_id=int(selected_row["レコードID"]),
            )
            update_started_at = perf_counter()
            if is_normal_change:
                operation = "update_schedule_change"
            else:
                operation = "update_temporary_schedule"

            try:
                if is_normal_change:
                    conn.execute(
                        """
                        UPDATE T_ScheduleChange
                        SET
                            ChangeTypeID = ?,
                            NewDoctorID = ?,
                            ChangeDetail = ?,
                            Reason = ?,
                            ChangedBy = ?,
                            Rpt2Flag = ?,
                            ActiveFlag = ?
                        WHERE ChangeID = ?
                        """,
                        (
                            int(edit_change_type_id),
                            edit_doctor_id,
                            edit_detail if edit_detail != "" else None,
                            edit_reason if edit_reason != "" else None,
                            edit_changed_by if edit_changed_by != "" else None,
                            1 if edit_visible_report2 else 0,
                            1 if edit_active else 0,
                            int(selected_row["レコードID"]),
                        ),
                    )
                else:
                    conn.execute(
                        """
                        UPDATE T_TemporarySchedule
                        SET
                            CalendarDate = ?,
                            TimeSlotID = ?,
                            Rpt1ClinDeptID = ?,
                            DoctorID = ?,
                            Room = ?,
                            Rpt1DisplayDoctorName = ?,
                            ChangeDetail = ?,
                            Reason = ?,
                            Rpt2Flag = ?,
                            ActiveFlag = ?
                        WHERE TempID = ?
                        """,
                        (
                            str(edit_date),
                            int(edit_timeslot_id),
                            int(edit_dept_id),
                            edit_doctor_id,
                            edit_room if edit_room != "" else None,
                            edit_rpt2_before_doctor if edit_rpt2_before_doctor != "" else None,
                            edit_detail if edit_detail != "" else None,
                            edit_reason if edit_reason != "" else None,
                            1 if edit_visible_report2 else 0,
                            1 if edit_active else 0,
                            int(selected_row["レコードID"]),
                        ),
                    )

                conn.commit()
                update_elapsed_ms = int((perf_counter() - update_started_at) * 1000)
                log_event(
                    "update_success",
                    "変更登録履歴検索",
                    request_id=update_request_id,
                    operation=operation,
                    elapsed_ms=update_elapsed_ms,
                )
                st.success("登録内容を更新しました。再検索して最新状態を確認してください。")
            except Exception as exc:
                log_event(
                    "update_failed",
                    "変更登録履歴検索",
                    request_id=update_request_id,
                    operation=operation,
                    error=type(exc).__name__,
                )
                raise

    st.markdown("---")
    st.subheader("変更届データ出力")
    st.caption("条件で検索し、対象行へチェックを入れると、下部プレビューおよびテンプレート反映Excel出力ができます。")

    with st.expander("絞り込み条件", expanded=True):
        filter_col1, filter_col2, filter_col3 = st.columns(3)
        with filter_col1:
            filter_dept = st.text_input("診療科名で検索", value="")
        with filter_col2:
            filter_doctor = st.text_input("変更後医師名で検索", value="")
        with filter_col3:
            filter_keyword = st.text_input("変更内容/備考で検索", value="")

    filtered_df = display_result_df.copy()
    if filter_dept:
        filtered_df = filtered_df[filtered_df["診療科名"].fillna("").str.contains(filter_dept, case=False, na=False)]
    if filter_doctor:
        filtered_df = filtered_df[filtered_df["変更後医師名"].fillna("").str.contains(filter_doctor, case=False, na=False)]
    if filter_keyword:
        keyword_mask = (
            filtered_df["変更内容"].fillna("").str.contains(filter_keyword, case=False, na=False)
            | filtered_df["備考"].fillna("").str.contains(filter_keyword, case=False, na=False)
        )
        filtered_df = filtered_df[keyword_mask]

    if filtered_df.empty:
        st.info("絞り込み条件に一致するデータがありません。")
    else:
        select_df = filtered_df.copy()
        select_df.insert(0, "出力対象", False)
        edited_selection = st.data_editor(
            select_df,
            hide_index=True,
            use_container_width=True,
            disabled=[col for col in select_df.columns if col != "出力対象"],
            column_config={"出力対象": st.column_config.CheckboxColumn("出力対象")},
            key="history_export_selection_editor",
        )

        selected_export_df = edited_selection[edited_selection["出力対象"]].drop(columns=["出力対象"])

        st.markdown("#### 出力対象プレビュー")
        if selected_export_df.empty:
            st.info("出力したい行にチェックを入れてください。")
        else:
            input_col1, input_col2 = st.columns(2)
            with input_col1:
                export_user = st.text_input("変更届出力者", value="")
            with input_col2:
                export_date = st.date_input("変更届出力日")

            selected_export_df = selected_export_df.copy()
            selected_export_df["変更届出力者"] = export_user if export_user else None
            selected_export_df["変更届出力日"] = str(export_date)
            st.dataframe(selected_export_df, use_container_width=True)

            st.markdown("#### テンプレート反映設定")
            st.info(
                "Excelテンプレート（.xlsx）にプレースホルダ（例: {{診療科#1}}）を入力しておくと、"
                "チェック済みの出力対象データを反映したExcelをダウンロードできます。"
            )
            st.caption(
                "プレースホルダ書式: {{列名#行番号}} / {{列名}} / {{固定:文字列}} "
                "（例: {{日付#1}}, {{曜日#1}}, {{診療科#2}}, {{変更届出力者#1}}）"
            )

            template_file = st.file_uploader(
                "変更届Excelテンプレート（.xlsx）",
                type=["xlsx"],
                key="history_template_uploader",
            )

            if template_file is not None:
                try:
                    template_bytes = template_file.getvalue()
                    sheet_title, replacements = build_history_placeholder_replacements(
                        template_bytes=template_bytes,
                        df=selected_export_df,
                    )
                    filled_excel = apply_replacements_preserve_media(
                        template_bytes=template_bytes,
                        sheet_title=sheet_title,
                        replacements=replacements,
                    )
                    download_clicked = st.download_button(
                        label="テンプレート反映版Excelダウンロード",
                        data=filled_excel,
                        file_name="変更届_テンプレート反映.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                    if download_clicked:
                        save_output_history(
                            conn=conn,
                            target_df=selected_export_df,
                            output_by=export_user,
                            output_date=str(export_date),
                        )
                        st.success("変更届出力履歴を登録しました。続けて別条件でも出力できます。")
                except Exception as exc:
                    st.error(f"テンプレートExcelへの反映に失敗しました: {exc}")
